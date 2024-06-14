import time
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import fb
import tweepy

# You can get a token on https://developers.facebook.com/tools/explorer
TOKEN_FB = 'YourToken'
# You can get a token on https://developer.twitter.com/en/portal/dashboard
consumer_key = 'Your api key'
consumer_secret = 'Your api key secret'
access_token = 'Your token'
access_token_secret = 'Your token secret'
wb = Workbook("Test.xlsx")
path_file = ("Test.xlsx")

def new_project():
    id_input = input("Enter the id of the proyect")
    name_input = input("Enter the name of the project")
    status_input = input("Enter the status of the project")
    delivery_input = input("Enter the delivery date of the project ")
    start_time = time.asctime()
    current_data2 = [id_input,
                     name_input,
                     status_input,
                     delivery_input,
                     start_time]
    print("[✓] Proyect Created.", "\n", "ID:", id_input, "\n"
          "Name:", name_input, "\n"
          "Status:", status_input, "\n"
          "Delivery date:", delivery_input, "\n"
          "Start time:", start_time, "\n")
    save_it = input("Do you want save it? y/n").strip().lower()
    if save_it == "y":
        save_progress()
    return current_data2

def save_progress(current_data2):
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    id = current_data2[0]
    name = current_data2[1]
    status = current_data2[2]
    delivery_date = current_data2[3]
    start_time = current_data2[4]
    datos = [id, name, status, delivery_date, start_time]
    sheet.append(datos)
    wb.save(path_file)

def finish_projects():
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    id_input = input("Enter ID project...")
    id = (int(id_input) + 1)
    finish_project_input = input("Are you sure you want to finalize project " + str(sheet.cell(row = id, column = 2).value) + "? y/n ").strip().lower()
    if finish_project_input == "y":
        sheet.cell(row = id, column = 3).value = "Terminado"
        print(sheet.cell(row = id, column = 3).value)
        horario_de_projecto_finalizado = time.asctime()
        sheet.cell(row = id, column = 6).value = horario_de_projecto_finalizado
        horario_excel = sheet.cell(id, column = 5).value
        horario_final = horario_de_projecto_finalizado
        horario_inicio = datetime.strptime(horario_excel,"%a %b %d %H:%M:%S %Y")
        horario_actual = datetime.strptime(horario_final, "%a %b %d %H:%M:%S %Y")
        duracion = horario_actual - horario_inicio
        duracion_en_horas = duracion.total_seconds() / 3600
        project_name = sheet.cell(id, column = 2).value
        print(f"Congratulations, the project {project_name} lasted {round(duracion_en_horas, 1)} hours.")
        sheet.cell(row = id, column = 7). value = round(duracion_en_horas, 1)
        wb.save("Test.xlsx")
        share_input =input("Do you want to share it? y/n").strip().lower()
        if share_input == "y":
            choice_input = input("""Where do you want to share it?
                                 1.Facebook
                                 2.Twitter/X
                                 3.Todos""")
            if choice_input == "1":
                share_facebook()
            elif choice_input == "2":
                share_twitter()
            elif choice_input == "3":
                share_facebook()
                share_twitter()
        else:
            print("Back to the menu!")
            return main()
    elif finish_project_input == "n":
        print("Back to the menu!")
        return main() 
           
def search_menu():
    print("""The proyects are...
    \t""")
    print("\n select an option:")
    option_input = input("""
    1. Edit project.
    2. Delete.
    3. Return main.
    """)
    if option_input == "1":
        edit()
    if option_input == "2":
        delete_input = input("Enter ID project...")
        delete_calculator = (int(delete_input) + 1)
        delete = input("Are you sure to delete it? y/n").strip().lower()
        if delete == "y":
            wb = load_workbook("Test.xlsx")
            sheet = wb["Projects"]
            sheet.delete_rows(delete_calculator)
            save_input = input("Do you want save it? y/n").strip().lower()
            if save_input == "y":
                wb.save("Test.xlsx")

def read_projects():
    print("""The proyects are...
    \t""")
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    for row in sheet.iter_rows():
        line = []
        for cell in row:
            line.append(str(cell.value))
        print("\t".join(line))

def edit():
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    id_input = input("Enter ID project...")
    id_calculator= (int(id_input) + 1)
    id = (id_calculator)
    edit_project_input = input("""What do you want to modify?
        1.Name project
        2.Status
        3.Delivery date.
        """)
    if edit_project_input == "1":
        edit_name_input = input("""What do you want to change it to?""")
        sheet.cell(row = id, column = 2).value = edit_name_input
        print(sheet.cell(row = id, column = 2).value)
        save_edit = input("Do you want save it? y/n").strip().lower()
        if save_edit == "y":
            wb.save("Test.xlsx")
        
    elif edit_project_input == "2":
        edit_status_input = input("""What do you want to change it to?""")
        sheet.cell(row = id, column = 3).value = edit_status_input
        print(sheet.cell(row = id, column = 3).value)
        save_edit = input("Do you want save it? y/n").strip().lower()
        if save_edit == "y":
            wb.save("Test.xlsx")

    elif edit_project_input == "3":
        edit_delivery_input = input("""What do you want to change it to?""")
        sheet.cell(row = id, column = 4).value = edit_delivery_input
        print(sheet.cell(row = id, column = 4).value)
        save_edit = input("Do you want save it? y/n").strip().lower()
        if save_edit == "y":
            wb.save("Test.xlsx")
            
def create_new_workbook():
    create_workbook_input = input("Create a new archive? y/n").lower()
    if create_workbook_input == "y":
        file_name_input = input("Perfect! What name do you want for the file?")
        confirm_workbook_input = input("Are you sure you want to give it this name " + file_name_input + "? y/n").strip().lower()
        if confirm_workbook_input == "y":
            create_workbook=(file_name_input)
            wb = Workbook()
            save_workbook = input("Do you want save it? y/n")
            if save_workbook == "y":
                print("Perfect! the file " + create_workbook + " has been saved")
                hoja = wb.active
                hoja.title = "Projects"
                sheet = wb["Projects"]
                columnas = ["ID", "Name", "Status", "Delivery date", "Start Time, Finish time, Project duration hours"]
                sheet.append(columnas)
                wb.save(create_workbook + ".xlsx")
                
def share_facebook():
    read_projects()
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    id_input = input("What project do you want to share?""ID")
    id_calculator= (int(id_input) + 1)
    id = (id_calculator)
    name_project = sheet.cell(row = id, column = 2).value
    horas_project = sheet.cell(row = id, column = 7).value
    """ Share the post on facebook """
    msg = f"I\'ve been working in the project {name_project} and it has taken me {horas_project} hours to finish it.'"
    facebook = fb.graph.api(TOKEN_FB)
    facebook.publish(cat = 'feed', id = 'me', message = msg)
    print('\n', msg)
    print('Shared! \n')

def share_twitter():
    read_projects()
    wb = load_workbook("Test.xlsx")
    sheet = wb["Projects"]
    id_input = input("What project do you want to share?""ID")
    id_calculator= (int(id_input) + 1)
    id = (id_calculator)
    name_project = sheet.cell(row = id, column = 2).value
    horas_project = sheet.cell(row = id, column = 7).value
    msg = f"I\'ve been working in the project {name_project} and it has taken me {horas_project} hours to finish it.'"
    client = tweepy.Client(consumer_key=consumer_key, consumer_secret=consumer_secret, access_token=access_token, access_token_secret=access_token_secret)
    client.create_tweet(text=msg)
    print('\n', msg)
    print('Shared! \n')

def share_post():
    choice_input = input("""Where do you want to share it?
                            1.Facebook
                            2.Twitter/X
                            3.Todos""")
    if choice_input == "1":
        share_facebook()
    elif choice_input == "2":
        share_twitter()
    elif choice_input == "3":
        share_facebook()
        share_twitter()

def contador():
    minutos_trabajados = 0
    comenzar_temporizador_input = input("Quieres trabajar en el proyecto? y/n").strip().lower()
    if comenzar_temporizador_input == "y":
        tiempo_input = input("Cuantos MINUTOS vas a dedicarle al proyecto hoy?")
        tiempo_segundos = tiempo_input * 60
        time.sleep(tiempo_segundos)
        while True:
            seguir_trabajando_input = input("¿Quieres seguir trabajando en el proyecto? (s/n): ").strip().lower()
            if seguir_trabajando_input == "s":
                tiempo_input = int(input("¿Cuántos MINUTOS vas a dedicarle al proyecto hoy? "))
                minutos_trabajados += tiempo_input
                print(f"¡Trabajando durante {tiempo_input} minutos!")
                print(f"Total de minutos trabajados hasta ahora: {minutos_trabajados} minutos")
                tiempo_segundos = tiempo_input * 60
                time.sleep(tiempo_segundos)
            elif seguir_trabajando_input == "n":
                print("¡Hasta luego!")
                break
            else:
                print("Respuesta no valida. Por favor, responde 's' para si o 'n' para no.")
def main():
    print("Welcome to Excel Simple Work Data")
    print("\nYour projects are:")

    while True:
        option = input("""
          +==========================================+
          |                                          |
          |               Work Project               |
          |                                          |
          +==========================================+
        
          [?] select an option:
        
          1. New project
          2. Search for a proyect
          3. Finish project
          4. Share
          5. Exit
          0. Options
          
        > """)
        if option == "1":
            new_project()

        if option == "2":
            read_projects()
            search_menu()

        if option == "3":
            finish_projects()
        
        if option == "4":
            share_post()
            
        if option == "5":
            print("Have a good day!")
            break

        if option == "0":
            create_new_workbook()

if __name__ == "__main__":
    main()